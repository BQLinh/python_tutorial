from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
import openpyxl

chrome_options = Options()
chrome_options.add_experimental_option("prefs", { 
    "profile.default_content_setting_values.notifications": 1 
})
chrome_options.add_experimental_option( "prefs",{'profile.managed_default_content_settings.javascript': 2})
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(10)

driver.get('https://hacom.vn/')

test = driver.find_element(By.ID, 'text-search')
test.send_keys('laptop')
test.send_keys(Keys.ENTER)
time.sleep(5)


# driver.find_element(By.CSS_SELECTOR, 'img.lazy.loaded')
list_laptop=driver.find_element(By.XPATH, '/html/body/div[6]/div[2]/div/div[2]/div/div[2]')
#print(list_laptop.text)
laptops = list_laptop.find_elements(By.CSS_SELECTOR, 'div.p-component.item')
mylist=[]
for product in laptops[:10]:
    name_elements = product.find_element(By.CLASS_NAME, 'p-name ')

    price_elements = product.find_element(By.CLASS_NAME, 'p-price')
    #print(price_elements.get_attribute('data-price'))

    code_elements = product.find_element(By.CLASS_NAME, 'p-sku')
    #print(code_elements.text)

    tag_link= product.find_element(By.TAG_NAME, 'a')
    link_elements=tag_link.get_attribute('href')
    #print(link_elements)

    img_link=product.find_element(By.TAG_NAME, 'img')
    img_elements=img_link.get_attribute('data-src')
    #print(img_elements)

    mylist.append({
        'name': name_elements.text,
        'cost': price_elements.get_attribute('data-price'),
        'code': code_elements.text,
        'link': link_elements,
        'img': img_elements
        })
    
# print(mylist)
# with open('data.txt', 'w', encoding='utf-8') as f:
my_file = openpyxl.load_workbook('test1.xlsx')
# new_file = openpyxl.Workbook()
ws = my_file.active

for i in range(len(mylist)):
    ws[f'A{i+1}'] = mylist[i]['name']
    ws[f'B{i+1}'] = mylist[i]['cost']
    ws[f'C{i+1}'] = mylist[i]['code']
    ws[f'D{i+1}'] = mylist[i]['link']
    ws[f'E{i+1}'] = mylist[i]['img']

my_file.save('test2.xlsx')